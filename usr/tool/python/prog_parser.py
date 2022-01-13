#!/usr/bin/env python3
# -*- coding: utf-8 -*-
## ============================================================================
## Copyright (c) 2021 Hsin-Hsien Yeh (Edward Yeh).
## All rights reserved.
## ----------------------------------------------------------------------------
## Filename         : prog_parser.py
## File Description : Programming Register Parser
## ----------------------------------------------------------------------------
## Author           : Edward Yeh
## Created On       : Sat 18 Dec 2021 11:48:41 PM CST
## Format           : Python module
## ----------------------------------------------------------------------------
## Reuse Issues     : 
## ----------------------------------------------------------------------------
## Release History  : 
##   2021/12/18 Edward Yeh : Initial version.
## ============================================================================

import os
import copy
import shutil
import argparse
import textwrap
import openpyxl

### Class Definition ###

class PatternList:
    """Programming pattern list"""

    def __init__(self, reg_fn: str, table_type: str, is_debug: bool):
    #{{{
        # reg_table = {addr1: reg_list1, addr2: reg_list2, ...}
        # reg_list = [type, tag, title, max_len, reg1, reg2, ...]
        # reg = [reg_name, msb, lsb, init_value, comment, row_idx]
        # pat_list = [[pat_name1, pat_table1], [pat_name2, pat_table2], ...]
        # pat_table = {addr1: val_list1, addr2: val_list2, ...}
        # val_list = [val1, val2, ...]

        self.is_debug = is_debug
        self.comment_sign = '#'
        self.reg_table = {}
        self.pat_list  = []

        if table_type == 'cfg':
            self.cfg_reg_parser(reg_fn)
        elif table_type == 'xls':
            self.xls_reg_parser(reg_fn)
        else:
            raise TypeError("Unsupport register table type ({})".format(table_type))
    #}}}

    def cfg_reg_parser(self, reg_fn: str):
        """Parse co nfig type register table"""  #{{{
        with open(reg_fn, 'r') as f:
            reg_list = [None, None, None, 0] 
            reg_act = False
            reg_addr = 0

            line = f.readline()
            while line:
                toks = line.split()
                if len(toks):
                    if toks[0] == 'H:' or toks[0] == 'A:' or toks[0] == 'T:':
                        if reg_act:
                            self.reg_table[reg_addr] = reg_list
                            reg_list = [None, None, None, 0]

                        if toks[0] == 'T:':
                            reg_list[1] = toks[1]
                            reg_act = False
                        else:
                            reg_addr = self.get_int_val(toks[1])
                            reg_list[0] = toks[0]
                            if len(toks) > 2:
                                reg_list[2] = ' '.join(toks[2:]).strip("\"\'")
                            reg_act = True
                    else:
                        reg = [toks[0].upper(), int(toks[1]), int(toks[2]), 
                               self.get_int_val(toks[3])]

                        if len(toks) > 4:
                            reg.append(' '.join(toks[4:]).strip("\"\'"))
                        else:
                            reg.append(None)

                        reg.append(None)  # row_idx is no use in this mode

                        name_len = len(reg[0])
                        if name_len > reg_list[3]:
                            reg_list[3] = name_len
                        reg_list.append(reg)

                line = f.readline()

            if reg_act:
                self.reg_table[reg_addr] = reg_list

        if self.is_debug:
            self.show_reg_table("=== REG TABLE PARSER ===")
    #}}}

    def xls_reg_parser(self, reg_fn: str):
        """Parse excel type register table"""  #{{{
        reg_list = [None, None, None, 0] 
        reg_act = False
        reg_addr = 0

        wb = openpyxl.load_workbook(reg_fn, data_only=True)
        ws = wb.worksheets[0]

        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
        for i in range(addr_col.index('ADDR')+1, len(addr_col)):
            row_idx = i + 1

            addr = addr_col[i]
            if addr is not None:
                if reg_act:
                    self.reg_table[reg_addr] = reg_list

                addr = str(addr)
                if addr == 'none':
                    break
                else:
                    reg_addr = int(addr, 16)
                    if ws.cell(row_idx, 1).font.__getattr__('color'):
                        reg_list = ['H:', None, None, 0]
                    else:
                        reg_list = ['A:', None, None, 0]
                    title = ws.cell(row_idx, 2).value
                    if title is not None:
                        reg_list[2] = str(title)

            reg_val = self.get_int_val(str(ws.cell(row_idx, 3).value))

            bits = str(ws.cell(row_idx, 4).value).split('_')
            if len(bits) > 1:
                msb, lsb = int(bits[0]), int(bits[1])
            else:
                msb = lsb = int(bits[0])

            toks = str(ws.cell(row_idx, 5).value).split('\n');
            reg_name = toks[0].upper()

            if len(toks) == 1:
                comment = None
            else:
                for i in range(1, len(toks)):
                    toks[i] = toks[i].strip()
                comment = ', '.join(toks[1:])

            name_len = len(reg_name)
            if name_len > reg_list[3]:
                reg_list[3] = name_len

            reg_list.append([reg_name, msb, lsb, reg_val, comment, row_idx])
            reg_act = True

        wb.close()

        if self.is_debug:
            self.show_reg_table("=== XLS TABLE PARSER ===")
    #}}}

    def ini_parser(self, ini_fn: str, is_batch=False, start=0, end=0):
        """Pattern parser for INI format"""  #{{{
        cfg_fns = []
        if is_batch:
            with open(ini_fn, 'r') as f:
                tmp_fns = f.readlines()

            if start == 0:
                start = 1
                end = len(tmp_fns)

            for i in range(start-1, end):
                cfg_fns.append(tmp_fns[i].rstrip())
        else:
            cfg_fns.append(ini_fn)

        for cfg_fn in cfg_fns:
            cfg = {}
            with open(cfg_fn, 'r') as f:
                line = f.readline()
                while line:
                    if line.startswith('['):
                        pass
                    elif line.startswith(self.comment_sign):
                        pass
                    else:
                        toks = line.split()
                        if len(toks) and toks[1] == '=':
                            cfg[toks[0].upper()] = self.get_int_val(toks[2])
                    line = f.readline()

            if self.is_debug:
                print("=== INI READ ({}) ===".format(cfg_fn))
                for item in cfg.items():
                    print(item)
                print()

            pat_table = {}
            for addr, reg_list in self.reg_table.items():
                val_list = []
                for reg in reg_list[4:]:
                    if reg_list[0] == 'H:':
                        val_list.append(reg[3])
                    else:
                        val_list.append(cfg.get(reg[0], reg[3])) 
                pat_table[addr] = val_list

            if self.is_debug:
                self.show_pat_content(pat_table, "=== PAT CONTENT ({}) ===".format(cfg_fn))

            pat_name = os.path.basename(cfg_fn)
            pat_name = os.path.splitext(pat_name)[0]
            self.pat_list.append([pat_name, pat_table])
    #}}}

    def hex_parser(self, hex_fn: str, is_batch=False, start=0, end=0):
        """Pattern parser for HEX format"""  #{{{
        cfg_fns = []
        if is_batch:
            with open(hex_fn, 'r') as f:
                tmp_fns = f.readlines()

            if start == 0:
                start = 1
                end = len(tmp_fns)

            for i in range(start-1, end):
                cfg_fns.append(tmp_fns[i].rstrip())
        else:
            cfg_fns.append(hex_fn)

        for cfg_fn in cfg_fns:
            cfg = {}
            with open(cfg_fn, 'r') as f:
                line = f.readline()
                while line:
                    addr = int(line[0:4], 16)
                    val = int(line[4:12], 16)
                    cfg[addr] = val
                    line = f.readline()

            if self.is_debug:
                print("=== HEX READ ({}) ===".format(cfg_fn))
                for key, val in cfg.items():
                    print("{:#06x}: {:#010x}".format(key, val))
                print()

            pat_table = {}
            for addr, reg_list in self.reg_table.items():
                val_list = []
                if addr in cfg:
                    for reg in reg_list[4:]:
                        if reg_list[0] == 'H:':
                            val_list.append(reg[3])
                        else:
                            mask = (1 << (reg[1] - reg[2] + 1)) - 1
                            reg_val = (cfg[addr] >> reg[2]) & mask
                            val_list.append(reg_val)
                else:
                    for reg in reg_list[4:]:
                        val_list.append(reg[3])
                pat_table[addr] = val_list

            if self.is_debug:
                self.show_pat_content(pat_table, "=== PAT CONTENT ({}) ===".format(cfg_fn))

            pat_name = os.path.basename(cfg_fn)
            pat_name = os.path.splitext(pat_name)[0]
            self.pat_list.append([pat_name, pat_table])
    #}}}

    def xls_parser(self, xls_fn: str, is_batch=False, start=0, end=0):
        """Pattern parser for INI format"""  #{{{
        wb = openpyxl.load_workbook(xls_fn, data_only=True)
        ws = wb.worksheets[0]
        
        if is_batch:
            if start == 0:
                start = 6
                end = ws.max_column
        else:
            start = end = 6

        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
        name_col = tuple(ws.iter_cols(5, 5, None, None, True))[0]
        row_st = addr_col.index('ADDR') + 1
        row_ed = addr_col.index('none')

        for j in range(start, end+1):
            val_col = tuple(ws.iter_cols(j, j, None, None, True))[0]
            pat_name = str(val_col[1])
            cfg = {}
            for i in range(row_st, row_ed):
                reg_name = name_col[i].split('\n')[0]
                if reg_name.upper() != 'RESERVED':
                    # val = self.get_int_val(str(val_col[i]))
                    val = int(str(val_col[i]), 16)
                    cfg[reg_name] = val

            if self.is_debug:
                print("=== XLS READ ({}) ===".format(pat_name))
                for item in cfg.items():
                    print(item)
                print()

            pat_table = {}
            for addr, reg_list in self.reg_table.items():
                val_list = []
                for reg in reg_list[4:]:
                    if reg_list[0] == 'H:':
                        val_list.append(reg[3])
                    else:
                        val_list.append(cfg.get(reg[0], reg[3]))
                pat_table[addr] = val_list

            if self.is_debug:
                self.show_pat_content(pat_table, "=== PAT CONTENT ({}) ===".format(pat_name))
                exit(0)

            self.pat_list.append([pat_name, pat_table])

        wb.close()
    #}}}

    def ini_dump(self):
        """Dump pattern with ini format""" #{{{
        if os.path.isdir("pat_out"):
            shutil.rmtree("pat_out")
        os.mkdir("pat_out")

        for pat in self.pat_list:
            is_first = True
            pat_fn = os.path.join("pat_out", pat[0]+".ini")
            with open(pat_fn, 'w') as f:
                for addr, reg_list in self.reg_table.items():
                    if reg_list[1] is not None:
                        if is_first:
                            f.write("{}\n".format(reg_list[1]))
                        else:
                            f.write("\n{}\n".format(reg_list[1]))

                    if reg_list[0] != 'H:':
                        for val, reg in zip(pat[1][addr], reg_list[4:]):
                            f.write("{} = {}".format(reg[0].lower(), val))
                            if reg[4] is not None:
                                f.write('{}{}\n'.format(' '*8, reg[4]))
                            else:
                                f.write("\n")
                        is_first = False
    #}}}

    def hex_dump(self):
        """Dump pattern with hex format""" #{{{
        if os.path.isdir("pat_out"):
            shutil.rmtree("pat_out")
        os.mkdir("pat_out")

        for pat in self.pat_list:
            pat_fn = os.path.join("pat_out", pat[0]+".dat")
            with open(pat_fn, 'w') as f:
                addr_list = sorted(tuple(pat[1].keys()))
                for addr in range(0, addr_list[-1]+4, 4):
                    word_val = 0
                    if addr in pat[1]:
                        for val, reg in zip(pat[1][addr], self.reg_table[addr][4:]):
                            word_val += val << reg[2]
                    f.write("{:04x}{:08x}\n".format(addr, word_val))
    #}}}

    def xls_dump(self, reg_fn : str):
        """Dump pattern with excel format""" #{{{
        if os.path.isdir("pat_out"):
            shutil.rmtree("pat_out")
        os.mkdir("pat_out")

        wb = openpyxl.load_workbook(reg_fn)
        ws = wb.worksheets[0]
        pat_idx = ws.max_column + 1

        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
        row_st = addr_col.index('ADDR') + 2
        row_ed = addr_col.index('none') + 1

        for i in range(row_st, row_ed):
            string = ws.cell(i, 5).value
            reg_name = string.split('\n')[0]
            if reg_name.upper() == 'RESERVED':
                row = ws.row_dimensions[i]
                cell = ws.cell(i, pat_idx, 0)
                cell.font = copy.copy(row.font)
                cell.fill = copy.copy(row.fill)
                cell.border = copy.copy(row.border)
                cell.alignment = copy.copy(row.alignment)

        for pat in self.pat_list:
            for addr, reg_list in self.reg_table.items():
                for val, reg in zip(pat[1][addr], reg_list[4:]):
                    row = ws.row_dimensions[reg[5]]
                    cell = ws.cell(reg[5], pat_idx, hex(val).upper()[2:])
                    cell.font = copy.copy(row.font)
                    cell.fill = copy.copy(row.fill)
                    cell.border = copy.copy(row.border)
                    cell.alignment = copy.copy(row.alignment)

            row = ws.row_dimensions[1]
            cell = ws.cell(1, pat_idx, pat_idx)
            cell.font = copy.copy(row.font)
            cell.fill = copy.copy(row.fill)
            cell.border = copy.copy(row.border)
            cell.alignment = copy.copy(row.alignment)

            row = ws.row_dimensions[2]
            cell = ws.cell(2, pat_idx, pat[0])
            cell.font = copy.copy(row.font)
            cell.fill = copy.copy(row.fill)
            cell.border = copy.copy(row.border)
            cell.alignment = copy.copy(row.alignment)

            pat_idx += 1

        wb.save(os.path.join("pat_out", "age_reg.xlsx"))
        wb.close()
    #}}}

    def get_int_val(self, str_val: str) -> int:
        """Convert string to integer (with HEX check)""" #{{{
        if str_val.startswith('0x') or str_val.startswith('0X') :
            return int(str_val, 16)
        else:
            return int(str_val)
    #}}}

    def show_reg_table(self, comment : str):
        """Show register table""" #{{{
        print(comment)
        for addr, reg_list in self.reg_table.items():
            print("{}".format([hex(addr)] + reg_list[0:4]))
            for reg in reg_list[4:]:
                print("  {}".format(reg))
        print()
    #}}}

    def show_pat_content(self, pat : dict, comment : str):
        """Show pattern content""" #{{{
        print(comment)
        for addr, val_list in pat.items():
            print("addr: {}".format(hex(addr)))
            print("{}".format(val_list))
        print()
    #}}}

def main(is_debug=False):
    """Main function""" #{{{
    parser = argparse.ArgumentParser(
            usage='%(prog)s [options]',
            formatter_class=argparse.RawDescriptionHelpFormatter,
            description=textwrap.dedent('''
                Programming Register Parser.

                Examples:
                  %(prog)s -creg reg_table.txt cfg hex age_reg.pat      
                  ## config in, hex out, with config format register table.
                '''))

    parser.add_argument('-c', dest='creg_fn', metavar='REG_FN', 
                              help='the register file with config format')
    parser.add_argument('-x', dest='xreg_fn', metavar='REG_FN', 
                              help='the register file with excel format')
    parser.add_argument('-b', dest='is_batch', action='store_true', 
                              help='enable batch mode (for cfg/hex in)')
    parser.add_argument('-s', dest='start_id', metavar='PAT_ID', type=int, default=0,
                              help='start pattern ID.')
    parser.add_argument('-e', dest='end_id', metavar='PAT_ID', type=int, default=0,
                              help='end pattern ID')
    parser.add_argument('in_fmt', help='input format (option: cfg/hex/xls)') 
    parser.add_argument('out_fmt', help='output format (option: cfg/hex/xls)') 
    parser.add_argument('pat_fn', help='pattern file name') 

    args = parser.parse_args()

    ## Parser register table
    pat_list = None
    if args.creg_fn != None:
        pat_list = PatternList(args.creg_fn, 'cfg', is_debug)
    elif args.xreg_fn != None:
        pat_list = PatternList(args.xreg_fn, 'xls', is_debug)
    else:
        print("[ERR] Register table unexisted (need to set -c or -x).")
        exit(1)

    ## Parse input pattern
    if args.in_fmt == 'cfg':
        pat_list.ini_parser(args.pat_fn, args.is_batch, args.start_id, args.end_id) 
    elif args.in_fmt == 'hex':
        pat_list.hex_parser(args.pat_fn, args.is_batch, args.start_id, args.end_id)
    elif args.in_fmt == 'xls':
        pat_list.xls_parser(args.pat_fn, args.is_batch, args.start_id, args.end_id)
    else:
        print("[ERR] Unsupport input format ({}).".format(args.in_fmt))
        exit(1)

    ## Dump pattern
    if args.out_fmt == 'cfg':
        pat_list.ini_dump()
    elif args.out_fmt == 'hex':
        pat_list.hex_dump()
    elif args.out_fmt == 'xls':
        pat_list.xls_dump(args.xreg_fn)
    else:
        print("[ERR] Unsupport output format ({}).".format(args.out_fmt))
        exit(1)
#}}}

if __name__ == '__main__':
    main(False)
else:
    pass
