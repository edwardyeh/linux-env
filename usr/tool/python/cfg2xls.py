#!/usr/bin/env python3
# -*- coding: utf-8 -*-
## +FHDR=======================================================================
## Copyright (c) 2022 Hsin-Hsien Yeh (Edward Yeh).
## All rights reserved.
## ----------------------------------------------------------------------------
## Filename         : cfg2xls.py
## File Description : Programming reference table convertor
## ----------------------------------------------------------------------------
## Author           : Edward Yeh
## Created On       : Wed 12 Jan 2022 02:34:24 AM CST
## Format           : Python module
## ----------------------------------------------------------------------------
## Reuse Issues     : 
## ----------------------------------------------------------------------------
## Release History  : 
## -FHDR=======================================================================

import os
import shutil
import argparse
import textwrap
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Side

### Class Definition ###

class RegisterTable:
    """Programming register table"""

    def __init__(self, reg_fn: str, table_type: str, is_debug: bool):
    #{{{
        # reg_table = {addr1: reg_list1, addr2: reg_list2, ...}
        # reg_list = [type, tag, title, max_len, reg1, reg2, ...]
        # reg = [reg_name, msb, lsb, init_value, comment, row_idx]

        self.is_debug = is_debug
        self.comment_sign = '#'
        self.reg_table = {}

        if table_type == 'cfg':
            self.cfg_reg_parser(reg_fn)
        elif table_type == 'xls':
            self.xls_reg_parser(reg_fn)
        else:
            raise TypeError("Unsupport register table type ({})".format(table_type))
    #}}}

    def cfg_reg_parser(self, reg_fn: str):
        """Parse co nfig type register table"""  #{{{
        with open(reg_fn, 'r') as f:
            reg_list = [None, None, None, 0] 
            reg_act = False
            reg_addr = 0

            line = f.readline()
            while line:
                toks = line.split()
                if len(toks):
                    if toks[0] == 'H:' or toks[0] == 'A:' or toks[0] == 'T:':
                        if reg_act:
                            self.reg_table[reg_addr] = reg_list
                            reg_list = [None, None, None, 0]

                        if toks[0] == 'T:':
                            reg_list[1] = toks[1]
                            reg_act = False
                        else:
                            reg_addr = self.get_int_val(toks[1])
                            reg_list[0] = toks[0]
                            if len(toks) > 2:
                                reg_list[2] = ' '.join(toks[2:]).strip("\"\'")
                            reg_act = True
                    else:
                        reg = [toks[0].upper(), int(toks[1]), int(toks[2]), 
                               self.get_int_val(toks[3])]

                        if len(toks) > 4:
                            reg.append(' '.join(toks[4:]).strip("\"\'"))
                        else:
                            reg.append(None)

                        reg.append(None)  # row_idx is no use in this mode

                        name_len = len(reg[0])
                        if name_len > reg_list[3]:
                            reg_list[3] = name_len
                        reg_list.append(reg)

                line = f.readline()

            if reg_act:
                self.reg_table[reg_addr] = reg_list

        if self.is_debug:
            self.show_reg_table("=== REG TABLE PARSER ===")
    #}}}

    def xls_reg_parser(self, reg_fn: str):
        """Parse excel type register table"""  #{{{
        reg_list = [None, None, None, 0] 
        reg_act = False
        reg_addr = 0

        wb = openpyxl.load_workbook(reg_fn, data_only=True)
        ws = wb.worksheets[0]

        addr_col = tuple(ws.iter_cols(1, 1, None, None, True))[0]
        for i in range(addr_col.index('ADDR')+1, len(addr_col)):
            row_idx = i + 1

            addr = addr_col[i]
            if addr is not None:
                if reg_act:
                    self.reg_table[reg_addr] = reg_list

                addr = str(addr)
                if addr == 'none':
                    break
                else:
                    reg_addr = int(addr, 16)
                    if ws.cell(row_idx, 1).font.__getattr__('color'):
                        reg_list = ['H:', None, None, 0]
                    else:
                        reg_list = ['A:', None, None, 0]
                    title = ws.cell(row_idx, 2).value
                    if title is not None:
                        reg_list[2] = str(title)

            reg_val = self.get_int_val(str(ws.cell(row_idx, 3).value))

            bits = str(ws.cell(row_idx, 4).value).split('_')
            if len(bits) > 1:
                msb, lsb = int(bits[0]), int(bits[1])
            else:
                msb = lsb = int(bits[0])

            toks = str(ws.cell(row_idx, 5).value).split('\n');
            reg_name = toks[0].upper()

            if len(toks) == 1:
                comment = None
            else:
                for i in range(1, len(toks)):
                    toks[i] = toks[i].strip()
                comment = ', '.join(toks[1:])

            name_len = len(reg_name)
            if name_len > reg_list[3]:
                reg_list[3] = name_len

            reg_list.append([reg_name, msb, lsb, reg_val, comment, row_idx])
            reg_act = True

        wb.close()

        if self.is_debug:
            self.show_reg_table("=== XLS TABLE PARSER ===")
    #}}}

    def cfg_dump(self, is_export: bool):
        """Dump config text""" #{{{
        with open('table_dump.txt', 'w') as f:
            is_first = True

            for addr, reg_list in self.reg_table.items():
                name_len = (reg_list[3] >> 2 << 2) + 4

                if not is_first:
                    f.write('\n')

                if reg_list[1] is not None:
                    f.write('T: {reg_list[1]}\n')

                f.write('{} {:9}'.format(reg_list[0], hex(addr)))

                if reg_list[2] is not None:
                    f.write(f'"{reg_list[2]}"\n')
                else:
                    f.write('\n')

                for reg in reg_list[4:]:
                    f.write('{}{}{:<4}{:<4}{:<12}'.format(reg[0].lower(), 
                                                          (' ' * (name_len - len(reg[0]))),
                                                          reg[1], 
                                                          reg[2], 
                                                          hex(reg[3])))

                    if reg[4] is not None:
                        f.write(f'"{reg[4]}"\n')
                    else:
                        f.write('\n')

                is_first = False
    #}}}

    def xls_dump(self, is_export: bool):
        """Dump excel file""" #{{{
        GREY_FONT = Font(color='808080')

        GREEN_FILL = PatternFill(fill_type='solid', start_color='92d050')
        GREY_FILL = PatternFill(fill_type='solid', start_color='dddddd')
        ORANGE_FILL = PatternFill(fill_type='solid', start_color='ffcc99')
        YELLOW_FILL = PatternFill(fill_type='solid', start_color='ffffcc')
        VIOLET_FILL = PatternFill(fill_type='solid', start_color='e6ccff')

        THIN_SIDE = Side(border_style='thin', color='000000')
        OUTER_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

        LT_ALIGN = Alignment(horizontal='left', vertical='top', wrapText=True)
        LC_ALIGN = Alignment(horizontal='left', vertical='center', wrapText=True)
        CT_ALIGN = Alignment(horizontal='center', vertical='top', wrapText=True)
        CC_ALIGN = Alignment(horizontal='center', vertical='center', wrapText=True)
        RT_ALIGN = Alignment(horizontal='right', vertical='top', wrapText=True)
        RC_ALIGN = Alignment(horizontal='right', vertical='center', wrapText=True)

        # Initial workbook 

        wb = openpyxl.Workbook()
        ws = wb.worksheets[0]

        ws.column_dimensions['A'].width = 15.46
        ws.column_dimensions['B'].width = 41.23
        ws.column_dimensions['C'].width = 10.31
        ws.column_dimensions['D'].width = 10.31
        ws.column_dimensions['E'].width = 41.23

        for i in (1, 2):
            row = ws.row_dimensions[i]
            row.fill = VIOLET_FILL
            row.border = OUTER_BORDER
            row.alignment = CC_ALIGN

        for i in (3, 4, 5):
            row = ws.row_dimensions[i]
            row.border = OUTER_BORDER
            row.alignment = CC_ALIGN

        row = ws.row_dimensions[6]
        row.fill = GREEN_FILL
        row.border = OUTER_BORDER
        row.alignment = CC_ALIGN

        for row in ws['A1:D5']:
            for cell in row:
                cell.border = OUTER_BORDER
                cell.alignment = CC_ALIGN

        values = ['Chip', 'Eng.', 'Date']
        for row, val in enumerate(values, start=2):
            cell = ws.cell(row, 1, val)
            cell.fill = ORANGE_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

        values = ['Number', 'FileName', 'Metion1', 'Metion2', 'PatternStatus']
        for row, val in enumerate(values, start=1):
            cell = ws.cell(row, 5, val)
            cell.fill = YELLOW_FILL
            cell.border = OUTER_BORDER
            cell.alignment = LC_ALIGN

        cell = ws.cell(6, 1, 'ADDR')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = CC_ALIGN

        cell = ws.cell(6, 2, 'Register')
        cell.fill = GREEN_FILL 
        cell.border = OUTER_BORDER
        cell.alignment = LC_ALIGN

        cell = ws.cell(6, 3, 'INI')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = RC_ALIGN

        cell = ws.cell(6, 4, 'Bits')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = RC_ALIGN

        cell = ws.cell(6, 5, 'Member')
        cell.fill = GREEN_FILL
        cell.border = OUTER_BORDER
        cell.alignment = LC_ALIGN

        if is_export:
            cell = ws.cell(1, 6, 6)
            cell.fill = VIOLET_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

            cell = ws.cell(2, 6, 'PAT-1')
            cell.fill = VIOLET_FILL
            cell.border = OUTER_BORDER
            cell.alignment = CC_ALIGN

        # Dump register

        row_st = row_ed = 7

        for addr in sorted(tuple(self.reg_table.keys())):
            reg_list = self.reg_table[addr]
            cell_font = GREY_FONT if reg_list[0] == 'H:' else Font()
            is_first = True

            for reg in reg_list[4:]:
                cell_fill = YELLOW_FILL if is_first else PatternFill()

                row = ws.row_dimensions[row_ed]
                row.font = cell_font
                row.fill = cell_fill
                row.border = OUTER_BORDER
                row.alignment = CC_ALIGN

                if is_first:
                    cell = ws.cell(row_ed, 1, hex(addr))
                    cell.font = cell_font
                    cell.fill = cell_fill
                    cell.border = OUTER_BORDER
                    cell.alignment = CT_ALIGN

                    cell = ws.cell(row_ed, 2, reg_list[2])
                    cell.font = cell_font
                    cell.fill = cell_fill
                    cell.border = OUTER_BORDER
                    cell.alignment = LT_ALIGN

                cell = ws.cell(row_ed, 3, hex(reg[3]))
                cell.font = cell_font
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = RT_ALIGN

                if reg[1] == reg[2]:
                    cell = ws.cell(row_ed, 4, str(reg[1]))
                else:
                    cell = ws.cell(row_ed, 4, '_'.join([str(reg[1]), str(reg[2])]))

                cell.font = cell_font
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = RT_ALIGN

                if reg[0] == 'RESERVED':
                    cell = ws.cell(row_ed, 5, reg[0].lower())
                else:
                    toks = [] if reg[4] is None else reg[4].split(',')
                    members = [reg[0]]

                    for tok in toks:
                        members.append(tok.strip())

                    cell = ws.cell(row_ed, 5, '\n'.join(members))

                cell.font = cell_font
                cell.fill = cell_fill
                cell.border = OUTER_BORDER
                cell.alignment = LT_ALIGN

                if is_export:
                    cell = ws.cell(row_ed, 6, hex(reg[3]).upper()[2:])
                    cell.font = cell_font
                    cell.fill = cell_fill
                    cell.border = OUTER_BORDER
                    cell.alignment = CC_ALIGN

                row_ed += 1
                is_first = False

        row = ws.row_dimensions[row_ed]
        row.fill = GREY_FILL
        row.border = OUTER_BORDER
        row.alignment = CC_ALIGN

        cell = ws.cell(row_ed, 1, 'none')
        cell.fill = GREY_FILL
        cell.border = OUTER_BORDER
        cell.alignment = CC_ALIGN
        row_ed += 1

        wb.save("table_dump.xlsx")
        wb.close()
    #}}}

    def get_int_val(self, str_val: str) -> int:
        """Convert string to integer (with HEX check)""" #{{{
        if str_val.startswith('0x') or str_val.startswith('0X') :
            return int(str_val, 16)
        else:
            return int(str_val)
    #}}}

    def show_reg_table(self, comment : str):
        """Show register table""" #{{{
        print(comment)
        for addr, reg_list in self.reg_table.items():
            print("{}".format([hex(addr)] + reg_list[0:4]))
            for reg in reg_list[4:]:
                print("  {}".format(reg))
        print()
    #}}}

### Main Function ###

def main(is_debug=False):
    """Main function""" #{{{
    parser = argparse.ArgumentParser(
            usage='%(prog)s [options]',
            formatter_class=argparse.RawDescriptionHelpFormatter,
            description=textwrap.dedent('''
                Programming Register Table converter.

                  config text to excel is default mode.
                '''))

    parser.add_argument('-i', dest='is_inverse', action='store_true', 
                              help='inverse mode (excel to config text)')
    parser.add_argument('-e', dest='is_export', action='store_true', 
                              help='export default setting')
    parser.add_argument('file', help='file name') 

    args = parser.parse_args()

    ## Parser register table & dump
    if args.is_inverse:
        pat_list = RegisterTable(args.file, 'xls', is_debug)
        pat_list.cfg_dump(args.is_export)
    else:
        pat_list = RegisterTable(args.file, 'cfg', is_debug)
        pat_list.xls_dump(args.is_export)
#}}}

if __name__ == '__main__':
    main(False)
else:
    pass
